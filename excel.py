'''
Working with excel table is implemented here by the class ExcelWB
'''

from openpyxl import load_workbook

class ExcelWB:
    '''
    Read and write an excel workbook with a given path.

    It's expected to have either one or more tables in the workbook:
      1. CheckList (required)
      2...n. Any table except FoundActs
    If FoundActs-table exists, it is raising an exception.

    Fields of CheckList-table:
      A2...An --  LastName
      B2...Bn -- FirstName
      C2...Cn -- Patronymic
    The first row of each field (column) is a name of that field.

    Methods defined here:
    1. read_list -- reading information from the table 'CheckList'
    2. write_acts -- creating the table 'FoundActs', writing found information
    into it.
    '''
    def __init__(self, file='actual.xlsx'):
        self.path = './Xlsx/' + file

    def read_list(self):
        '''
        Reading the data from table CheckList. The four columns are expected:
        LastName, FirstName, Patronymic, BirthDate. Names of these fields are not
        important except the LastName field, it should take a place at A1 cell,
        other names should be at the first row, too.
        '''

        try:
            sheet = load_workbook(filename = self.path)['CheckList']

            res_list = []

            for row in sheet.values:
                if row[0]:
                    if row[0] != 'LastName':
                        ln, fn, pat = row[0], row[1], row[2]

                        res_list.append((ln, fn, pat))
                    else:
                        continue
                else:
                    break

            return res_list

        except Exception:
            raise IOError('Either he workbook {} doesn\'t exist or the table \
            "CheckList" doesn\'t exist or it\'s not correct '.format(self.path))


    def write_acts(self, found_acts_dict):
        '''
        Write collected information.
        into the <name>.xlsx file.
        The file mustn't have a table 'FoundActs' else an exception raises.
        Argument -- a dictionary where keys are names and values are lists of tuples
        with an information from the site.
        '''
        try:
            workbook = load_workbook(filename = self.path)

        except IOError:
            raise IOError('Probably the file {} doesn\'t exist'.format(self.path))


        try:
            sheet = workbook['FoundActs']

        except KeyError:
            sheet = workbook.create_sheet(title='FoundActs')

        else:
            raise IOError('Probably the workbook is used already, the new \
            one shouldn\'t have the FoundActs-sheet')


        sheet['A1'] = 'ФИО'
        sheet['B1'] = 'Суд'
        sheet['C1'] = 'Номер дела'
        sheet['D1'] = 'Дата поступления'
        sheet['E1'] = 'Информация по делу'
        sheet['F1'] = 'Судья'
        sheet['G1'] = 'Результат слушания'
        sheet['H1'] = 'Судебные акты'

        index_list = []
        is_first = True
        for name, info_list in found_acts_dict.items():
            if not is_first:
                start_index = index_list[-1][2] + 1
            else:
                start_index = 2
                is_first = False

            fin_index = start_index + len(info_list)
            index_list.append((name, start_index, fin_index))

        for item in index_list:

            sheet.merge_cells(start_row=item[1],
                              start_column=1,
                              end_row=item[2],
                              end_column=1)

            sheet.cell(column=1, row=item[1], value=item[0])

            for (record, row) in zip(found_acts_dict[item[0]],
                                         range(item[1], item[2])):
                for col in range(2, 9):
                    sheet.cell(column=col, row=row, value=record[col-2])

        workbook.save(filename = self.path)
